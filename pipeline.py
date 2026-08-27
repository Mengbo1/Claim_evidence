from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .extractors import extract_facts
from .linker import generate_links
from .scoring import aggregate_claim_assessments
from .workbook_io import load_source_rows, read_claims, write_coded_workbook


_PACKAGE_PARENT = Path(__file__).resolve().parents[1]
_PORTABLE_PROJECT = _PACKAGE_PARENT.parent if _PACKAGE_PARENT.name.lower() == "code" else None

if _PORTABLE_PROJECT is not None and (_PORTABLE_PROJECT / "data").exists():
    DEFAULT_INPUT = _PORTABLE_PROJECT / "data" / "claim_evidence_input.xlsx"
    DEFAULT_OUTPUT = _PORTABLE_PROJECT / "results" / "claim_evidence_coded_results.xlsx"
else:
    # Backward-compatible defaults for the original workspace. Explicit input
    # and output arguments continue to override these paths.
    DEFAULT_INPUT = Path(r"C:\Users\lenovo\Documents\Malware report\outputs\mlware_standardized_review\mlware_minimal_coding_v2.xlsx")
    DEFAULT_OUTPUT = Path(r"C:\Users\lenovo\Documents\Malware report\outputs\mlware_standardized_review\mlware_coding_simplified.xlsx")


def run_pipeline(input_path: Path = DEFAULT_INPUT, output_path: Path = DEFAULT_OUTPUT) -> dict[str, str | int]:
    source_workbook = load_workbook(input_path, read_only=True, data_only=True)
    evidence_rows, evidence_index, claim_rows, claim_index = load_source_rows(source_workbook)
    report_to_sample = {
        str(row[evidence_index["report_id"]] or "").strip(): str(row[evidence_index["sample_id"]] or "").strip()
        for row in evidence_rows[1:]
    }
    claims = read_claims(claim_rows, claim_index, report_to_sample)
    synthetic_claims = []
    if "synthetic claims" in source_workbook.sheetnames:
        synthetic_rows = [
            list(row) for row in source_workbook["synthetic claims"].iter_rows(values_only=True)
        ]
        if synthetic_rows:
            synthetic_index = {
                str(value or "").strip(): position
                for position, value in enumerate(synthetic_rows[0])
            }
            synthetic_claims = read_claims(
                synthetic_rows, synthetic_index, report_to_sample, "synthetic claims",
            )
    source_workbook.close()
    facts = extract_facts(evidence_rows[1:], evidence_index)
    links = generate_links(claims, facts)
    assessments = aggregate_claim_assessments(claims, links)
    synthetic_links = generate_links(synthetic_claims, facts)
    synthetic_assessments = aggregate_claim_assessments(synthetic_claims, synthetic_links)
    write_coded_workbook(
        input_path, output_path, facts, links, assessments,
        synthetic_assessments if synthetic_claims else None,
    )
    return {
        "input": str(input_path), "output": str(output_path),
        "claims": len(claims), "facts": len(facts), "links": len(links),
        "synthetic_claims": len(synthetic_claims),
        "synthetic_links": len(synthetic_links),
    }
