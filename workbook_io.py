from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from .models import Claim, Fact, Link
from .text_utils import headers_to_index, normalize, text


FACT_HEADERS = [
    "fact_id", "evidence_id", "sample_id", "report_id", "evidence_basis",
    "evidence_domain", "evidence_predicate", "evidence_object", "object_type",
    "event_status", "specificity",
]
LINK_HEADERS = ["link_id", "claim_id", "fact_id", "match_type"]


def load_source_rows(workbook) -> tuple[list[list[Any]], dict[str, int], list[list[Any]], dict[str, int]]:
    evidence_rows = [list(row) for row in workbook["Evidence Inventory"].iter_rows(values_only=True)]
    claim_rows = [list(row) for row in workbook["Claim Assessment"].iter_rows(values_only=True)]
    return evidence_rows, headers_to_index(evidence_rows[0]), claim_rows, headers_to_index(claim_rows[0])


def read_claims(
    claim_rows: list[list[Any]], index: dict[str, int], report_to_sample: dict[str, str],
    sheet_name: str = "Claim Assessment",
) -> list[Claim]:
    claim_id_header = "claim_id" if "claim_id" in index else "synthetic_claim_id"
    required = {claim_id_header, "source_report_ids", "claim_type", "claim_predicate", "claim_object", "claim_text"}
    missing = sorted(required - set(index))
    if missing:
        raise ValueError(f"{sheet_name} is missing columns: {', '.join(missing)}")

    claims: list[Claim] = []
    for row in claim_rows[1:]:
        report_ids = tuple(
            part.strip() for part in text(row[index["source_report_ids"]]).replace("/", ";").split(";") if part.strip()
        )
        # Older workbooks use claim_subject (for example, sample001) instead
        # of sample_id. They are equivalent for this project, so accept either.
        sample_column = "sample_id" if "sample_id" in index else "claim_subject"
        sample_id = text(row[index[sample_column]]) if sample_column in index else report_to_sample.get(report_ids[0], "")
        if not sample_id:
            raise ValueError(f"Cannot resolve a sample_id for claim {text(row[index[claim_id_header]])}.")
        claims.append(Claim(
            claim_id=text(row[index[claim_id_header]]),
            sample_id=sample_id,
            report_ids=report_ids,
            claim_type=text(row[index["claim_type"]]),
            predicate=text(row[index["claim_predicate"]]),
            object=text(row[index["claim_object"]]),
            text=text(row[index["claim_text"]]),
        ))
    return claims


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    for column_cells in sheet.columns:
        letter = column_cells[0].column_letter
        width = min(max(len(text(cell.value)) for cell in column_cells) + 2, 45)
        sheet.column_dimensions[letter].width = max(12, width)


def _replace_table(sheet, name: str) -> None:
    for table_name in list(sheet.tables.keys()):
        del sheet.tables[table_name]
    last_cell = sheet.cell(sheet.max_row, sheet.max_column).coordinate
    table = Table(displayName=name, ref=f"A1:{last_cell}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    sheet.add_table(table)


def _write_generated_sheet(workbook, sheet_name: str, headers: list[str], rows: list[list[Any]], table_name: str) -> None:
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    _style_sheet(sheet)
    _replace_table(sheet, table_name)


def _ensure_claim_assessment_columns(
    sheet, assessments: dict[str, str], table_name: str = "ClaimAssessmentTable",
) -> None:
    headers = [text(cell.value) for cell in sheet[1]]
    if "assessment" in headers and "assessment_human" not in headers:
        sheet.cell(1, headers.index("assessment") + 1).value = "assessment_human"
        headers[headers.index("assessment")] = "assessment_human"

    for header in ("assessment_auto", "auto_human_agree"):
        if header not in headers:
            headers.append(header)
            sheet.cell(1, len(headers)).value = header

    index = headers_to_index(headers)
    claim_id_header = "claim_id" if "claim_id" in index else "synthetic_claim_id"
    if claim_id_header not in index:
        raise ValueError(f"{sheet.title} is missing a claim identifier column.")
    for row_number in range(2, sheet.max_row + 1):
        claim_id = text(sheet.cell(row_number, index[claim_id_header] + 1).value)
        if not claim_id:
            continue
        auto_value = assessments.get(claim_id, "not_verifiable")
        sheet.cell(row_number, index["assessment_auto"] + 1).value = auto_value
        human_value = text(sheet.cell(row_number, index["assessment_human"] + 1).value)
        agreement = "pending_human" if not human_value else (
            "agree" if normalize(human_value) == normalize(auto_value) else "disagree"
        )
        sheet.cell(row_number, index["auto_human_agree"] + 1).value = agreement

    _style_sheet(sheet)
    _replace_table(sheet, table_name)


def write_coded_workbook(
    input_path: Path, output_path: Path, facts: list[Fact], links: list[Link],
    assessments: dict[str, str], synthetic_assessments: dict[str, str] | None = None,
) -> None:
    workbook = load_workbook(input_path)

    # Coding Results duplicated claim-level assessment; Claim Assessment is the
    # single source of truth after the new automatic columns are added there.
    for sheet_name in ("Evidence Facts", "Claim Evidence Links", "Coding Results", "Coding Guide"):
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    _ensure_claim_assessment_columns(workbook["Claim Assessment"], assessments)
    if synthetic_assessments is not None and "synthetic claims" in workbook.sheetnames:
        _ensure_claim_assessment_columns(
            workbook["synthetic claims"], synthetic_assessments, "SyntheticClaimsTable",
        )
    # The final workbook shows only facts that actually participate in at least
    # one claim–evidence link. The full original evidence remains untouched in
    # Evidence Inventory, so no source material is lost.
    used_fact_ids = {link.fact_id for link in links}
    used_facts = [fact for fact in facts if fact.fact_id in used_fact_ids]
    _write_generated_sheet(workbook, "Evidence Facts", FACT_HEADERS, [fact.as_row() for fact in used_facts], "EvidenceFactsTable")
    _write_generated_sheet(workbook, "Claim Evidence Links", LINK_HEADERS, [link.as_row() for link in links], "ClaimEvidenceLinksTable")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
